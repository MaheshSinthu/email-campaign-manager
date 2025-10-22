import csv
import io
import openpyxl
from django.core.validators import validate_email
from django.core.exceptions import ValidationError

def parse_csv(file_obj):
    """
    Accepts a file-like object for CSV and yields dicts: {'name':..., 'email':...}
    """
    decoded = file_obj.read().decode('utf-8')
    reader = csv.DictReader(io.StringIO(decoded))
    for row in reader:
        yield {'name': row.get('name') or '', 'email': row.get('email')}

def parse_excel(file_obj):
    wb = openpyxl.load_workbook(file_obj)
    ws = wb.active
    headers = [str(cell.value).strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        yield {'name': data.get('name') or '', 'email': data.get('email')}

def validate_and_clean_rows(rows):
    """
    rows: iterable of dicts {'name','email'}
    returns: tuple(valid_list, invalid_list) where invalid_list items contain reason
    """
    valid = []
    invalid = []
    seen = set()
    for r in rows:
        email = (r.get('email') or '').strip()
        name = (r.get('name') or '').strip()
        if not email:
            invalid.append({'row': r, 'reason': 'Missing email'})
            continue
        try:
            validate_email(email)
        except ValidationError:
            invalid.append({'row': r, 'reason': 'Invalid email format'})
            continue
        if email in seen:
            invalid.append({'row': r, 'reason': 'Duplicate in file'})
            continue
        seen.add(email)
        valid.append({'name': name, 'email': email})
    return valid, invalid
